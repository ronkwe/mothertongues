from openpyxl import load_workbook
import pandas as pd
from mtd.parsers.utils import BaseParser
from mtd.exceptions import UnsupportedFiletypeError
from mtd.parsers.utils import ResourceManifest

class Parser(BaseParser):
    '''
    Parse data for MTD **TODO: test worksheet location. Skipheader in manifest skips first row. Location in manifest decides worksheet.

    :param ResourceManifest manifest: Manifest for parser
    :param str resource_path: path to file 
    '''
    def __init__(self, manifest: ResourceManifest, resource_path: str):
        self.manifest = manifest
        try:
            work_book = load_workbook(resource_path)
            if "location" in self.manifest:
                work_sheet = work_book["location"]
            else:
                work_sheet = work_book.active
            if self.manifest['skipheader']:
                min_row = 2
            else:
                min_row = 1
            self.resource = work_sheet.iter_rows(min_row=min_row)
        except:
            raise UnsupportedFiletypeError(resource_path)
        
        self.entry_template = self.manifest['targets']

    def getCellValue(self, entry, col):
        for c in entry:
            if c.column == col:
                return c.value
        return ''

    def resolve_targets(self):
        word_list = []
        for entry in self.resource:
            word_list.append(self.fill_entry_template(self.entry_template, entry, self.getCellValue))
        return word_list

    def parse(self):
        try:
            data = self.resolve_targets()
            return {"manifest": self.manifest, "data": pd.DataFrame(data)}
        except:
            print('no targets')